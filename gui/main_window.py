"""
Cửa sổ chính của ứng dụng PyQt5
"""
import sys
import os
import sip
import cv2
import numpy as np
import pandas as pd
import openpyxl
import shutil
from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QTabWidget,
    QLineEdit, QSpinBox, QTableWidget, QTableWidgetItem,
    QMessageBox, QFileDialog, QProgressBar, QCheckBox, QDateEdit,
    QHeaderView, QAbstractItemView
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QDate, QPropertyAnimation, QEasingCurve, QRect
from PyQt5.QtGui import QImage, QPixmap, QFont
from .account_manager import AccountManagerWidget
from .student_manager import StudentManagerWidget
from .login_dialog import LoginDialog
from datetime import datetime
from .camera_thread import CameraThread, CollectionCameraThread
from .simple_recognition_thread import SimpleRecognitionThread  # Nhận diện đơn giản
from .multi_recognition_thread import MultiRecognitionThread  # Nhận diện nhiều người
# from .recognition_thread import RecognitionThread  # Không dùng nữa
# from .training_thread import TrainingThread  # Không dùng nữa
from utils import (
    FaceDetector,
    load_attendance_records,
    get_person_list,
    upsert_student_info,
    get_student_info,
    save_student_encoding,
    load_student_encodings,
    student_id_exists
)
try:
    from utils import FaceEncoder
except Exception:
    FaceEncoder = None
from config import CAMERA_INDEX, DATASET_DIR, IMAGE_SIZE, ATTENDANCE_DATE_FORMAT, DUPLICATE_SIMILARITY_THRESHOLD
import pandas as pd
import time


class MainWindow(QMainWindow):
    """
    Cửa sổ chính của ứng dụng
    """
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hệ Thống Điểm Danh Khuôn Mặt")
        self.setGeometry(100, 100, 1200, 800)
        
        # Khởi tạo biến
        self.camera_thread = None
        self.collection_thread = None
        self.recognition_thread = None
        self.multi_recognition_thread = None
        self.training_thread = None
        self.is_camera_running = False
        self.face_detector = None
        self.collected_images = 0
        self.current_person_name = ""
        self.current_student_id = None
        self.current_class = None
        self.current_folder_name = None
        self.current_frame = None
        self.attendance_today = []
        self._last_fps_ts = None
        self._fps = 0.0
        self.recognition_status_label = None
        self.fps_label = None
        self.collection_progress = None
        self.auto_stop_checkbox = None
        self.collection_person_dir = None
        self.collection_created_dir = False
        self.collection_saved_any = False
        
        # Cấu hình tự động tắt camera sau khi điểm danh
        self.auto_stop_after_attendance = False  # Mặc định KHÔNG tự tắt camera sau khi điểm danh
        
        # Khởi tạo face detector
        try:
            self.face_detector = FaceDetector()
            print("✓ Face detector đã khởi tạo thành công")
        except Exception as e:
            print(f"✗ Không thể khởi tạo face detector: {e}")
            QMessageBox.warning(
                None, "Cảnh báo",
                f"Không thể khởi tạo face detector!\n\n"
                f"Lỗi: {str(e)}\n\n"
                f"Vui lòng kiểm tra:\n"
                f"1. File haarcascade_frontalface_default.xml đã tồn tại\n"
                f"2. Thư viện opencv-python đã được cài đặt"
            )
        
        # Tạo giao diện
        self.init_ui()
        self.apply_styles()
        # Cờ tránh mở hộp thoại đăng nhập 2 lần do sự kiện currentChanged bắn nhiều lần
        self._login_in_progress = False

    def show_toast(self, message, success=True, duration_ms=2500):
        """
        Hiển thị toast góc trên bên phải (xanh thành công, đỏ thất bại)
        """
        parent = self.centralWidget() if self.centralWidget() else self
        toast = QLabel(message, parent)
        toast.setObjectName("ToastMessage")
        bg = "#27ae60" if success else "#e74c3c"
        toast.setStyleSheet(
            f"""
            QLabel#ToastMessage {{
                background: {bg};
                color: white;
                padding: 10px 16px;
                border-radius: 10px;
                font-weight: 600;
            }}
            """
        )
        # Là widget con bên trong cửa sổ chính
        toast.setWindowFlags(Qt.FramelessWindowHint)
        toast.setAttribute(Qt.WA_ShowWithoutActivating)
        toast.setAttribute(Qt.WA_TransparentForMouseEvents)
        toast.adjustSize()
        margin = 16
        # Tính toạ độ theo hệ quy chiếu của widget cha
        end_x = parent.width() - toast.width() - margin
        end_y = margin
        # Bắt đầu từ ngoài rìa phải, trượt vào
        start_x = parent.width() + 40
        start_y = end_y
        # Đặt hình học ban đầu và hiển thị
        toast.setGeometry(QRect(start_x, start_y, toast.width(), toast.height()))
        toast.setWindowOpacity(1.0)
        toast.show()
        toast.raise_()
        # Animation trượt vào
        anim_in = QPropertyAnimation(toast, b"geometry", self)
        anim_in.setDuration(350)
        anim_in.setEasingCurve(QEasingCurve.OutCubic)
        anim_in.setStartValue(QRect(start_x, start_y, toast.width(), toast.height()))
        anim_in.setEndValue(QRect(end_x, end_y, toast.width(), toast.height()))
        toast._anim_in = anim_in
        anim_in.start()
        # Bắt đầu mờ dần và đóng sau duration
        def fade_and_close():
            anim_out = QPropertyAnimation(toast, b"windowOpacity", self)
            anim_out.setDuration(300)
            anim_out.setStartValue(1.0)
            anim_out.setEndValue(0.0)
            anim_out.setEasingCurve(QEasingCurve.OutCubic)
            # Xóa widget sau khi mờ
            def _cleanup():
                try:
                    toast.close()
                    toast.deleteLater()
                except Exception:
                    pass
            anim_out.finished.connect(_cleanup)
            toast._anim_out = anim_out
            anim_out.start()
        QTimer.singleShot(duration_ms, fade_and_close)
    
    def init_ui(self):
        """
        Khởi tạo giao diện
        """
        # Widget trung tâm
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout chính
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # Tiêu đề
        title_label = QLabel("HỆ THỐNG ĐIỂM DANH KHUÔN MẶT")
        title_label.setAlignment(Qt.AlignCenter)
        title_font = QFont("Arial", 20, QFont.Bold)
        title_label.setFont(title_font)
        title_label.setStyleSheet("color: #2c3e50; padding: 10px;")
        main_layout.addWidget(title_label)
        
        # Tab widget
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # Theo yêu cầu: Nhận diện & Điểm danh nhiều người không cần đăng nhập.
        # Khu vực quản trị (Thu thập dữ liệu, DS Điểm danh, Quản lý tài khoản) yêu cầu đăng nhập khi truy cập.
        self.create_recognition_tab()
        self.create_multi_recognition_tab()
        self.create_admin_portal_tab()
        # Theo dõi chuyển tab để tự bật đăng nhập khi vào Quản Trị
        try:
            self.tabs.currentChanged.connect(self._on_tab_changed)
        except Exception:
            pass
        
        # Status bar
        self.statusBar().showMessage("Sẵn sàng")
    
    def create_recognition_tab(self):
        """
        Tab nhận diện và điểm danh
        """
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Video display
        self.video_label = QLabel()
        self.video_label.setMinimumSize(640, 480)
        self.video_label.setAlignment(Qt.AlignCenter)
        self.video_label.setStyleSheet("border: 2px solid #3498db; background-color: #ecf0f1;")
        layout.addWidget(self.video_label)

        status_layout = QHBoxLayout()
        self.recognition_status_label = QLabel("Sẵn sàng")
        self.recognition_status_label.setObjectName("StatusPill")
        self.fps_label = QLabel("FPS: --")
        self.fps_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        status_layout.addWidget(self.recognition_status_label)
        status_layout.addStretch(1)
        status_layout.addWidget(self.fps_label)
        layout.addLayout(status_layout)

        # Buttons
        button_layout = QHBoxLayout()

        self.btn_start_camera = QPushButton("Bật Camera")
        self.btn_start_camera.clicked.connect(self.start_camera)
        self.btn_start_camera.setStyleSheet(self.get_button_style("#27ae60"))
        button_layout.addWidget(self.btn_start_camera)

        self.btn_stop_camera = QPushButton("Tắt Camera")
        self.btn_stop_camera.clicked.connect(self.stop_camera)
        self.btn_stop_camera.setEnabled(False)
        self.btn_stop_camera.setStyleSheet(self.get_button_style("#e74c3c"))
        button_layout.addWidget(self.btn_stop_camera)

        self.btn_reload_data = QPushButton("🔄 Reload Dữ Liệu")
        self.btn_reload_data.clicked.connect(self.reload_recognition_data)
        self.btn_reload_data.setEnabled(False)
        self.btn_reload_data.setStyleSheet(self.get_button_style("#f39c12"))
        self.btn_reload_data.setToolTip("Cập nhật dữ liệu người dùng mới mà không cần khởi động lại")
        button_layout.addWidget(self.btn_reload_data)

        self.auto_stop_checkbox = QCheckBox("Tự tắt sau khi điểm danh")
        self.auto_stop_checkbox.setChecked(self.auto_stop_after_attendance)
        button_layout.addWidget(self.auto_stop_checkbox)

        layout.addLayout(button_layout)

        # Bỏ nhật ký nhận diện (dùng toast thay thế)
        self.recognition_log = None

        tab.setLayout(layout)
        self.tabs.addTab(tab, "Nhận Diện & Điểm Danh")
    
    def create_collection_tab(self):
        """
        Tab thu thập dữ liệu
        """
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Input
        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("Tên sinh viên:"))
        self.input_person_name = QLineEdit()
        self.input_person_name.setPlaceholderText("Nhập tên sinh viên...")
        input_layout.addWidget(self.input_person_name)
        
        input_layout.addWidget(QLabel("Số ảnh:"))
        self.input_num_images = QSpinBox()
        self.input_num_images.setMinimum(1)  # Tối thiểu 1 ảnh
        self.input_num_images.setMaximum(100)
        self.input_num_images.setValue(1)  # Mặc định 1 ảnh
        input_layout.addWidget(self.input_num_images)

        # Thêm Mã SV và Lớp
        input_layout.addWidget(QLabel("Mã SV:"))
        self.input_student_id = QLineEdit()
        self.input_student_id.setPlaceholderText("VD: B21DCCN001")
        input_layout.addWidget(self.input_student_id)

        input_layout.addWidget(QLabel("Lớp:"))
        self.input_class = QLineEdit()
        self.input_class.setPlaceholderText("VD: D21CQCN01-N")
        input_layout.addWidget(self.input_class)
        
        layout.addLayout(input_layout)
        
        # Video display
        self.collection_video_label = QLabel()
        self.collection_video_label.setMinimumSize(640, 480)
        self.collection_video_label.setAlignment(Qt.AlignCenter)
        self.collection_video_label.setStyleSheet("border: 2px solid #3498db; background-color: #ecf0f1;")
        layout.addWidget(self.collection_video_label)

        # Buttons
        button_layout = QHBoxLayout()
        
        self.btn_start_collection = QPushButton("Bắt Đầu Thu Thập")
        self.btn_start_collection.clicked.connect(self.start_collection)
        self.btn_start_collection.setStyleSheet(self.get_button_style("#3498db"))
        button_layout.addWidget(self.btn_start_collection)
        
        self.btn_stop_collection = QPushButton("Dừng Thu Thập")
        self.btn_stop_collection.clicked.connect(self.stop_collection)
        self.btn_stop_collection.setEnabled(False)
        self.btn_stop_collection.setStyleSheet(self.get_button_style("#e74c3c"))
        button_layout.addWidget(self.btn_stop_collection)
        
        layout.addLayout(button_layout)

        self.collection_progress = QProgressBar()
        self.collection_progress.setMinimum(0)
        self.collection_progress.setMaximum(1)
        self.collection_progress.setValue(0)
        layout.addWidget(self.collection_progress)

        # Bỏ nhật ký tiến trình (dùng toast thay thế)
        self.collection_log = None
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Thu Thập Dữ Liệu")
    
    def create_training_tab(self):
        """
        Tab huấn luyện mô hình
        """
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Info
        info_label = QLabel("Huấn luyện mô hình nhận diện khuôn mặt từ dữ liệu đã thu thập")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Button
        self.btn_train = QPushButton("Bắt Đầu Huấn Luyện")
        self.btn_train.clicked.connect(self.start_training)
        self.btn_train.setStyleSheet(self.get_button_style("#9b59b6"))
        self.btn_train.setMinimumHeight(50)
        layout.addWidget(self.btn_train)
        
        # Log
        self.training_log = QTextEdit()
        self.training_log.setReadOnly(True)
        layout.addWidget(QLabel("Nhật ký huấn luyện:"))
        layout.addWidget(self.training_log)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Huấn Luyện Mô Hình")
    
    def create_attendance_tab(self):
        """
        Tab xem điểm danh
        """
        tab = QWidget()
        layout = QVBoxLayout()
        
        date_bar = QHBoxLayout()
        date_bar.addWidget(QLabel("Ngày:"))
        self.attendance_date_edit = QDateEdit()
        self.attendance_date_edit.setCalendarPopup(True)
        # Dùng định dạng hiển thị của Qt, còn định dạng truy vấn DB dùng ATTENDANCE_DATE_FORMAT (Python)
        self.attendance_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.attendance_date_edit.setDate(QDate.currentDate())
        date_bar.addWidget(self.attendance_date_edit)
        date_bar.addStretch(1)
        self.btn_refresh_attendance = QPushButton("Làm Mới")
        self.btn_refresh_attendance.clicked.connect(self.refresh_attendance)
        self.btn_refresh_attendance.setStyleSheet(self.get_button_style("#16a085"))
        date_bar.addWidget(self.btn_refresh_attendance)
        self.btn_export_attendance = QPushButton("Xuất File")
        self.btn_export_attendance.clicked.connect(self.export_attendance)
        self.btn_export_attendance.setStyleSheet(self.get_button_style("#2980b9"))
        date_bar.addWidget(self.btn_export_attendance)
        layout.addLayout(date_bar)
        # Thanh tìm kiếm
        search_bar = QHBoxLayout()
        search_bar.addWidget(QLabel("Tìm tên:"))
        self.search_name = QLineEdit()
        self.search_name.setPlaceholderText("VD: Nguyễn Văn A")
        search_bar.addWidget(self.search_name)
        search_bar.addWidget(QLabel("Mã SV:"))
        self.search_sid = QLineEdit()
        self.search_sid.setPlaceholderText("VD: B21DCCN001")
        search_bar.addWidget(self.search_sid)
        search_bar.addWidget(QLabel("Lớp:"))
        self.search_class = QLineEdit()
        self.search_class.setPlaceholderText("VD: D21CQCN01-N")
        search_bar.addWidget(self.search_class)
        self.btn_search_attendance = QPushButton("Tìm kiếm")
        self.btn_search_attendance.setStyleSheet(self.get_button_style("#7f8c8d"))
        self.btn_search_attendance.clicked.connect(self.refresh_attendance)
        search_bar.addWidget(self.btn_search_attendance)
        layout.addLayout(search_bar)
        
        # Table
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(6)
        self.attendance_table.setHorizontalHeaderLabels(["Mã SV", "Tên", "Lớp", "Ngày", "Thời gian", "Độ tin cậy"])
        # UX: bảng đẹp và dễ xem hơn
        self.attendance_table.setAlternatingRowColors(True)
        self.attendance_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.attendance_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.attendance_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.attendance_table.setSortingEnabled(True)
        self.attendance_table.setShowGrid(False)
        header = self.attendance_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        vheader = self.attendance_table.verticalHeader()
        if vheader:
            vheader.setVisible(False)
        layout.addWidget(self.attendance_table)
        
        # Summary
        self.attendance_summary = QLabel()
        self.attendance_summary.setStyleSheet("padding: 10px; background-color: #ecf0f1;")
        layout.addWidget(self.attendance_summary)
        
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Danh Sách Điểm Danh")

    def create_multi_recognition_tab(self):
        """
        Tab nhận diện và điểm danh nhiều người cùng lúc
        """
        tab = QWidget()
        layout = QVBoxLayout()

        # Video display
        self.multi_video_label = QLabel()
        self.multi_video_label.setMinimumSize(640, 480)
        self.multi_video_label.setAlignment(Qt.AlignCenter)
        self.multi_video_label.setStyleSheet("border: 2px solid #8e44ad; background-color: #ecf0f1;")
        layout.addWidget(self.multi_video_label)

        # Buttons
        button_layout = QHBoxLayout()

        self.btn_start_multi = QPushButton("Bật Multi-Cam")
        self.btn_start_multi.clicked.connect(self.start_multi_recognition)
        self.btn_start_multi.setStyleSheet(self.get_button_style("#8e44ad"))
        button_layout.addWidget(self.btn_start_multi)

        self.btn_stop_multi = QPushButton("Tắt Multi-Cam")
        self.btn_stop_multi.clicked.connect(self.stop_multi_recognition)
        self.btn_stop_multi.setEnabled(False)
        self.btn_stop_multi.setStyleSheet(self.get_button_style("#e74c3c"))
        button_layout.addWidget(self.btn_stop_multi)

        self.btn_reload_multi = QPushButton("🔄 Reload Dữ Liệu")
        self.btn_reload_multi.clicked.connect(self.reload_multi_data)
        self.btn_reload_multi.setEnabled(False)
        self.btn_reload_multi.setStyleSheet(self.get_button_style("#f39c12"))
        self.btn_reload_multi.setToolTip("Cập nhật dữ liệu người dùng cho chế độ nhiều người")
        button_layout.addWidget(self.btn_reload_multi)

        layout.addLayout(button_layout)

        tab.setLayout(layout)
        self.tabs.addTab(tab, "Điểm Danh Nhiều Người")

    def create_admin_portal_tab(self):
        """Tạo tab 'Quản Trị' (sẽ tự bật đăng nhập khi người dùng chuyển tới tab này)."""
        self.admin_tab = QWidget()
        layout = QVBoxLayout()
        self.admin_tab.setLayout(layout)
        # Để trống; nội dung sẽ được mount sau khi đăng nhập thành công
        self.tabs.addTab(self.admin_tab, "Quản Trị")
        self._admin_portal_mounted = False

    def _on_tab_changed(self, idx):
        try:
            widget = self.tabs.widget(idx)
        except Exception:
            return
        # Nếu vào tab Quản Trị và chưa mount nội dung -> yêu cầu đăng nhập
        if hasattr(self, 'admin_tab') and widget is self.admin_tab and not getattr(self, '_admin_portal_mounted', False):
            if getattr(self, '_login_in_progress', False):
                return
            self._login_in_progress = True
            dlg = LoginDialog(self)
            result = dlg.exec_()
            # Hộp thoại đăng nhập sẽ tự đóng sau khi accept/reject
            if result == dlg.Accepted:
                uname, role = dlg.current_user()
                self.current_user = uname
                self.current_role = role or 'user'
                # Đảm bảo hộp thoại đã đóng hoàn toàn
                dlg.close()
                dlg.deleteLater()
                
                # Cập nhật UI ngay để dialog biến mất
                from PyQt5.QtWidgets import QApplication
                QApplication.processEvents()
                
                # Sau khi đăng nhập: chuyển toàn bộ giao diện sang các tab quản trị theo yêu cầu
                self._switch_to_management_tabs()
                self._admin_portal_mounted = True
                
                # Tự động chuyển sang tab đầu tiên của phần quản trị và refresh UI
                try:
                    if self.tabs.count() > 0:
                        self.tabs.setCurrentIndex(0)
                    # Cập nhật UI ngay lập tức để hiển thị các tab mới
                    QApplication.processEvents()
                    self.repaint()
                except Exception:
                    pass
            else:
                # Quay lại tab đầu nếu hủy đăng nhập
                try:
                    self.tabs.setCurrentIndex(0)
                except Exception:
                    pass
            self._login_in_progress = False

    def _mount_admin_portal(self, tab_widget):
        """Thay nội dung tab quản trị bằng QTabWidget con sau khi đăng nhập."""
        old_layout = tab_widget.layout()
        if old_layout:
            while old_layout.count():
                item = old_layout.takeAt(0)
                w = item.widget()
                if w:
                    w.setParent(None)
        inner = QTabWidget()
        role = getattr(self, 'current_role', 'user') or 'user'
        if role in ('admin', 'collector'):
            inner.addTab(self._build_collection_widget(), "Thu Thập Dữ Liệu")
        if role in ('admin', 'attendance_viewer'):
            inner.addTab(self._build_attendance_widget(), "Danh Sách Điểm Danh")
        if role == 'admin':
            inner.addTab(AccountManagerWidget(self, current_user_role=role), "Quản Lý Tài Khoản")
        new_layout = QVBoxLayout()
        tab_widget.setLayout(new_layout)
        new_layout.addWidget(inner)

    def _clear_all_tabs(self):
        try:
            while self.tabs.count() > 0:
                w = self.tabs.widget(0)
                self.tabs.removeTab(0)
                if w:
                    w.setParent(None)
        except Exception:
            pass

    def _switch_to_management_tabs(self):
        """Thay toàn bộ tabs bằng 3 tab: Danh Sách Điểm Danh, Quản Lý Người Dùng, Thu Thập Dữ Liệu."""
        self._clear_all_tabs()
        
        # Xóa toolbar cũ nếu có
        try:
            main_widget = self.centralWidget()
            main_layout = main_widget.layout()
            if main_layout and hasattr(self, 'logout_btn'):
                for i in range(main_layout.count()):
                    item = main_layout.itemAt(i)
                    if item and item.layout():
                        layout = item.layout()
                        if layout.indexOf(self.logout_btn) >= 0:
                            self.logout_btn.setParent(None)
                            main_layout.removeItem(item)
                            item.layout().deleteLater()
                            break
        except Exception:
            pass
        
        # Thêm thanh toolbar với nút thoát
        toolbar_layout = QHBoxLayout()
        toolbar_layout.addStretch(1)
        self.logout_btn = QPushButton("🚪 Thoát Quản Trị")
        self.logout_btn.setStyleSheet(self.get_button_style("#e74c3c"))
        self.logout_btn.clicked.connect(self._logout_from_admin)
        toolbar_layout.addWidget(self.logout_btn)
        
        # Thêm toolbar vào layout chính
        main_widget = self.centralWidget()
        main_layout = main_widget.layout()
        if main_layout:
            # Chèn toolbar vào trước tabs (sau title, trước tabs)
            main_layout.insertLayout(1, toolbar_layout)
        
        role = getattr(self, 'current_role', 'admin') or 'admin'
        
        # 1) Danh Sách Điểm Danh - cho admin và attendance_viewer
        if role in ('admin', 'attendance_viewer'):
            self.tabs.addTab(self._build_attendance_widget(), "Danh Sách Điểm Danh")
        
        # 2) Quản Lý Người Dùng - CHỈ admin
        if role == 'admin':
            self.tabs.addTab(AccountManagerWidget(self, current_user_role=role), "Quản Lý Người Dùng")
        
        # 3) Quản Lý Danh Sách Sinh Viên - cho admin và data_manager
        if role in ('admin', 'data_manager'):
            self.tabs.addTab(StudentManagerWidget(self, current_user_role=role), "Quản Lý Danh Sách Sinh Viên")
        
        # 4) Thu Thập Dữ Liệu - cho admin và data_manager
        if role in ('admin', 'data_manager'):
            self.tabs.addTab(self._build_collection_widget(), "Thu Thập Dữ Liệu")
        
        # Cập nhật UI ngay sau khi tạo tabs
        try:
            from PyQt5.QtWidgets import QApplication
            QApplication.processEvents()
            self.statusBar().showMessage("Đã vào chế độ quản trị")
        except Exception:
            pass
    
    def _logout_from_admin(self):
        """Thoát khỏi chế độ quản trị và quay lại giao diện ban đầu."""
        reply = QMessageBox.question(
            self, "Xác nhận thoát", "Bạn có muốn thoát khỏi chế độ quản trị không?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self._restore_default_tabs()
    
    def _restore_default_tabs(self):
        """Khôi phục lại các tab mặc định (trước khi đăng nhập)."""
        # Xóa toolbar logout nếu có
        try:
            main_widget = self.centralWidget()
            main_layout = main_widget.layout()
            if main_layout and hasattr(self, 'logout_btn'):
                # Tìm và xóa layout chứa nút logout
                for i in range(main_layout.count()):
                    item = main_layout.itemAt(i)
                    if item and item.layout():
                        layout = item.layout()
                        if layout.indexOf(self.logout_btn) >= 0:
                            # Xóa widget trước
                            self.logout_btn.setParent(None)
                            # Xóa layout
                            main_layout.removeItem(item)
                            item.layout().deleteLater()
                            break
        except Exception:
            pass
        
        # Xóa tất cả tabs
        self._clear_all_tabs()
        
        # Khôi phục tabs ban đầu
        self.create_recognition_tab()
        self.create_multi_recognition_tab()
        self.create_admin_portal_tab()
        
        # Reset trạng thái đăng nhập
        self._admin_portal_mounted = False
        if hasattr(self, 'current_user'):
            delattr(self, 'current_user')
        if hasattr(self, 'current_role'):
            delattr(self, 'current_role')
        
        try:
            self.statusBar().showMessage("Đã thoát khỏi chế độ quản trị")
        except Exception:
            pass

    def _build_collection_widget(self):
        tab = QWidget()
        layout = QVBoxLayout()
        tab.setLayout(layout)
        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("Tên sinh viên:"))
        self.input_person_name = QLineEdit()
        self.input_person_name.setPlaceholderText("Nhập tên sinh viên...")
        input_layout.addWidget(self.input_person_name)
        input_layout.addWidget(QLabel("Số ảnh:"))
        self.input_num_images = QSpinBox()
        self.input_num_images.setMinimum(1)
        self.input_num_images.setMaximum(100)
        self.input_num_images.setValue(1)
        input_layout.addWidget(self.input_num_images)
        input_layout.addWidget(QLabel("Mã SV:"))
        self.input_student_id = QLineEdit()
        self.input_student_id.setPlaceholderText("VD: B21DCCN001")
        input_layout.addWidget(self.input_student_id)
        input_layout.addWidget(QLabel("Lớp:"))
        self.input_class = QLineEdit()
        self.input_class.setPlaceholderText("VD: D21CQCN01-N")
        input_layout.addWidget(self.input_class)
        layout.addLayout(input_layout)
        self.collection_video_label = QLabel()
        self.collection_video_label.setMinimumSize(640, 480)
        self.collection_video_label.setAlignment(Qt.AlignCenter)
        self.collection_video_label.setStyleSheet("border: 2px solid #3498db; background-color: #ecf0f1;")
        layout.addWidget(self.collection_video_label)
        button_layout = QHBoxLayout()
        self.btn_start_collection = QPushButton("Bắt Đầu Thu Thập")
        self.btn_start_collection.clicked.connect(self.start_collection)
        self.btn_start_collection.setStyleSheet(self.get_button_style("#3498db"))
        button_layout.addWidget(self.btn_start_collection)
        self.btn_stop_collection = QPushButton("Tạm Dừng")
        self.btn_stop_collection.clicked.connect(self.pause_or_resume_collection)
        self.btn_stop_collection.setEnabled(False)
        self.btn_stop_collection.setStyleSheet(self.get_button_style("#f39c12"))
        button_layout.addWidget(self.btn_stop_collection)
        layout.addLayout(button_layout)
        self.collection_progress = QProgressBar()
        self.collection_progress.setMinimum(0)
        self.collection_progress.setMaximum(1)
        self.collection_progress.setValue(0)
        layout.addWidget(self.collection_progress)
        self.collection_log = None
        return tab

    def _build_attendance_widget(self):
        tab = QWidget()
        layout = QVBoxLayout()
        tab.setLayout(layout)
        date_bar = QHBoxLayout()
        date_bar.addWidget(QLabel("Ngày:"))
        self.attendance_date_edit = QDateEdit()
        self.attendance_date_edit.setCalendarPopup(True)
        self.attendance_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.attendance_date_edit.setDate(QDate.currentDate())
        date_bar.addWidget(self.attendance_date_edit)
        date_bar.addStretch(1)
        self.btn_refresh_attendance = QPushButton("Làm Mới")
        self.btn_refresh_attendance.clicked.connect(self.refresh_attendance)
        self.btn_refresh_attendance.setStyleSheet(self.get_button_style("#16a085"))
        date_bar.addWidget(self.btn_refresh_attendance)
        self.btn_export_attendance = QPushButton("Xuất File")
        self.btn_export_attendance.clicked.connect(self.export_attendance)
        self.btn_export_attendance.setStyleSheet(self.get_button_style("#2980b9"))
        date_bar.addWidget(self.btn_export_attendance)
        layout.addLayout(date_bar)
        search_bar = QHBoxLayout()
        search_bar.addWidget(QLabel("Tìm tên:"))
        self.search_name = QLineEdit()
        self.search_name.setPlaceholderText("VD: Nguyễn Văn A")
        search_bar.addWidget(self.search_name)
        search_bar.addWidget(QLabel("Mã SV:"))
        self.search_sid = QLineEdit()
        self.search_sid.setPlaceholderText("VD: B21DCCN001")
        search_bar.addWidget(self.search_sid)
        search_bar.addWidget(QLabel("Lớp:"))
        self.search_class = QLineEdit()
        self.search_class.setPlaceholderText("VD: D21CQCN01-N")
        search_bar.addWidget(self.search_class)
        self.btn_search_attendance = QPushButton("Tìm kiếm")
        self.btn_search_attendance.setStyleSheet(self.get_button_style("#7f8c8d"))
        self.btn_search_attendance.clicked.connect(self.refresh_attendance)
        search_bar.addWidget(self.btn_search_attendance)
        layout.addLayout(search_bar)
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(6)
        self.attendance_table.setHorizontalHeaderLabels(["Mã SV", "Tên", "Lớp", "Ngày", "Thời gian", "Độ tin cậy"])
        self.attendance_table.setAlternatingRowColors(True)
        self.attendance_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.attendance_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.attendance_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.attendance_table.setSortingEnabled(True)
        self.attendance_table.setShowGrid(False)
        header = self.attendance_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        vheader = self.attendance_table.verticalHeader()
        if vheader:
            vheader.setVisible(False)
        layout.addWidget(self.attendance_table)
        self.attendance_summary = QLabel()
        self.attendance_summary.setStyleSheet("padding: 10px; background-color: #ecf0f1;")
        layout.addWidget(self.attendance_summary)
        return tab

    def start_multi_recognition(self):
        """
        Bật chế độ nhận diện nhiều người
        """
        try:
            if not self.face_detector:
                self.face_detector = FaceDetector()
            self.multi_recognition_thread = MultiRecognitionThread(CAMERA_INDEX, self.face_detector)
            self.multi_recognition_thread.frame_ready.connect(self.update_multi_frame)
            self.multi_recognition_thread.attendance_marked.connect(self.on_multi_attendance_marked)
            self.multi_recognition_thread.error_occurred.connect(self.handle_camera_error)
            self.multi_recognition_thread.start()

            self.btn_start_multi.setEnabled(False)
            self.btn_stop_multi.setEnabled(True)
            self.btn_reload_multi.setEnabled(True)
            self.statusBar().showMessage("Chế độ điểm danh nhiều người đang chạy")
        except Exception as e:
            self.handle_camera_error(f"Lỗi khởi động Multi: {str(e)}")

    def stop_multi_recognition(self):
        """
        Tắt chế độ nhận diện nhiều người
        """
        if self.multi_recognition_thread:
            self.multi_recognition_thread.stop()
            self.multi_recognition_thread = None

        if hasattr(self, 'multi_video_label') and self.multi_video_label:
            self.multi_video_label.clear()
            self.multi_video_label.setText("Camera đã dừng")

        if hasattr(self, 'btn_start_multi'):
            self.btn_start_multi.setEnabled(True)
        if hasattr(self, 'btn_stop_multi'):
            self.btn_stop_multi.setEnabled(False)
        if hasattr(self, 'btn_reload_multi'):
            self.btn_reload_multi.setEnabled(False)
        self.statusBar().showMessage("Đã dừng chế độ nhiều người")

    def update_multi_frame(self, qt_image):
        """
        Cập nhật frame cho tab nhiều người
        """
        try:
            # Kiểm tra xem widget có tồn tại và chưa bị xóa chưa
            if not hasattr(self, 'multi_video_label') or self.multi_video_label is None:
                return
                
            # Kiểm tra xem widget có còn hợp lệ không
            if not sip.isdeleted(self.multi_video_label):
                pixmap = QPixmap.fromImage(qt_image)
                if not pixmap.isNull():
                    self.multi_video_label.setPixmap(pixmap.scaled(
                        self.multi_video_label.size(), 
                        Qt.KeepAspectRatio, 
                        Qt.SmoothTransformation
                    ))
        except Exception as e:
            print(f"Lỗi khi cập nhật frame: {str(e)}")
            # Nếu có lỗi, dừng luồng nhận diện để tránh lỗi tiếp tục xảy ra
            if hasattr(self, 'multi_recognition_thread') and self.multi_recognition_thread:
                self.multi_recognition_thread.stop()
                self.multi_recognition_thread.wait()
                self.multi_recognition_thread = None

    def on_multi_attendance_marked(self, name, time_str):
        """
        Xử lý khi điểm danh thành công trong chế độ nhiều người
        """
        # Không dừng camera, chỉ hiển thị thông báo và refresh danh sách
        try:
            info = get_student_info(name)
            sid = info[0] if info and len(info) > 0 else ''
        except Exception:
            sid = ''
        display_name = f"{name}{(' - ' + sid) if sid else ''}"
        self.show_toast(f"{display_name} điểm danh thành công", success=True)
        self.statusBar().showMessage(f"Đã điểm danh (Multi) cho {display_name} lúc {time_str}")
        self.refresh_attendance()

    def reload_multi_data(self):
        """
        Reload dữ liệu nhận diện cho chế độ nhiều người
        """
        if not self.multi_recognition_thread:
            QMessageBox.warning(self, "Lỗi", "Chế độ nhiều người chưa khởi động!")
            return
        try:
            if hasattr(self.multi_recognition_thread, 'reload_templates'):
                self.multi_recognition_thread.reload_templates()
                QMessageBox.information(self, "Thành công", "Đã reload dữ liệu người dùng (Multi)")
            else:
                QMessageBox.warning(self, "Lỗi", "Thread nhiều người không hỗ trợ reload")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể reload dữ liệu: {str(e)}")
    
    def get_button_style(self, color):
        """
        Lấy style cho button
        """
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 10px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:disabled {{
                background-color: #95a5a6;
            }}
        """
    
    def darken_color(self, color):
        """
        Làm tối màu
        """
        # Simple darkening
        return color.replace("27ae60", "1e8449").replace("e74c3c", "c0392b") \
                   .replace("3498db", "2980b9").replace("9b59b6", "8e44ad") \
                   .replace("16a085", "138d75").replace("2980b9", "21618c")
    
    def start_camera(self):
        """
        Bật camera nhận diện và điểm danh
        """
        self.log_message("Đang khởi động hệ thống nhận diện...", self.recognition_log)
        
        try:
            # Sử dụng SimpleRecognitionThread - không cần huấn luyện
            self.recognition_thread = SimpleRecognitionThread(CAMERA_INDEX, self.face_detector)
            self.recognition_thread.frame_ready.connect(self.update_camera_frame)
            self.recognition_thread.face_recognized.connect(self.on_face_recognized)
            self.recognition_thread.attendance_marked.connect(self.on_attendance_marked)
            # Tự dừng khi thất bại nhận diện (đứng 3s nhưng không xác định danh tính)
            if hasattr(self.recognition_thread, 'recognition_failed'):
                self.recognition_thread.recognition_failed.connect(self.handle_recognition_failed)
            self.recognition_thread.error_occurred.connect(self.handle_camera_error)
            self.recognition_thread.start()
            
            self.is_camera_running = True
            self.btn_start_camera.setEnabled(False)
            self.btn_stop_camera.setEnabled(True)
            self.btn_reload_data.setEnabled(True)
            self.statusBar().showMessage("Hệ thống nhận diện đang chạy")
            self.log_message("✓ Hệ thống nhận diện đã khởi động", self.recognition_log)
            if self.auto_stop_checkbox:
                self.auto_stop_after_attendance = self.auto_stop_checkbox.isChecked()
        except Exception as e:
            self.handle_camera_error(f"Lỗi khởi động: {str(e)}")
    
    def stop_camera(self):
        """
        Tắt camera
        """
        if self.recognition_thread:
            self.recognition_thread.stop()
            self.recognition_thread = None
        
        if self.camera_thread:
            self.camera_thread.stop()
            self.camera_thread = None
        
        self.video_label.clear()
        self.video_label.setText("Camera đã dừng")
        
        self.is_camera_running = False
        self.btn_start_camera.setEnabled(True)
        self.btn_stop_camera.setEnabled(False)
        self.btn_reload_data.setEnabled(False)
        self.statusBar().showMessage("Camera đã dừng")
        self.log_message("Đã dừng hệ thống nhận diện", self.recognition_log)
    
    def _validate_student_id_unique(self, student_id, student_name):
        """
        Đảm bảo mã sinh viên chưa được dùng cho người khác.
        """
        cleaned_id = (student_id or "").strip()
        if not cleaned_id:
            return True
        exists, owner = student_id_exists(cleaned_id, exclude_name=student_name)
        if exists:
            warning = (
                f"Mã sinh viên '{cleaned_id}' đã được gán cho sinh viên '{owner}'.\n"
                "Vui lòng nhập mã sinh viên khác trước khi tiếp tục thu thập."
            )
            QMessageBox.warning(self, "Mã sinh viên trùng", warning)
            self.show_toast("Mã sinh viên bị trùng, vui lòng nhập mã khác", success=False)
            if hasattr(self, 'input_student_id') and self.input_student_id:
                self.input_student_id.setFocus()
                self.input_student_id.selectAll()
            return False
        return True
    
    def start_collection(self):
        """
        Bắt đầu thu thập dữ liệu
        """
        # Nếu đang thu thập và chỉ cần đổi tên
        if self.collection_thread and self.collection_thread.isRunning():
            person_name = self.input_person_name.text().strip()
            if not person_name:
                QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên sinh viên!")
                return
            
            # Cập nhật thông tin sinh viên (nếu có)
            student_id = (self.input_student_id.text() or "").strip()
            if not self._validate_student_id_unique(student_id, person_name):
                return
            # Chỉ reset thông tin người mới sau khi đã xác thực mã SV
            self.current_person_name = person_name
            class_name = (self.input_class.text() or "").strip()
            # Lưu vào biến trạng thái
            try:
                existing_sid, existing_class = get_student_info(self.current_person_name)
            except Exception:
                existing_sid, existing_class = (None, None)
            existing_sid = existing_sid.strip() if isinstance(existing_sid, str) else existing_sid
            existing_class = existing_class.strip() if isinstance(existing_class, str) else existing_class
            self.current_student_id = student_id or existing_sid
            self.current_class = class_name or existing_class
            self.collected_images = 0
            num_images = self.input_num_images.value()
            # Tạo thư mục cho người này: tên_masv
            suffix_id = (self.current_student_id or "unknown").replace(" ", "")
            self.current_folder_name = f"{self.current_person_name}_{suffix_id}"
            person_dir = os.path.join(DATASET_DIR, self.current_folder_name)
            folder_existed = os.path.exists(person_dir)
            os.makedirs(person_dir, exist_ok=True)
            self.collection_person_dir = person_dir
            self.collection_created_dir = not folder_existed
            self.collection_saved_any = False
            
            # Reload known encodings khi đổi người (sau khi đã có current_folder_name)
            self.load_known_encodings(exclude_name=self.current_folder_name)
            
            self.log_message(f"Bắt đầu thu thập {num_images} ảnh cho {person_name}", self.collection_log)
            return
        
        # Nếu chưa có thread, khởi tạo mới
        person_name = self.input_person_name.text().strip()
        if not person_name:
            QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên sinh viên!")
            return
        
        if not self.face_detector:
            QMessageBox.warning(self, "Lỗi", "Face detector chưa được khởi tạo!")
            return
        
        student_id = (self.input_student_id.text() or "").strip()
        if not self._validate_student_id_unique(student_id, person_name):
            return
        self.current_person_name = person_name
        self.collected_images = 0
        num_images = self.input_num_images.value()
        # Cập nhật thông tin sinh viên (nếu có)
        class_name = (self.input_class.text() or "").strip()
        # Lưu vào biến trạng thái
        try:
            existing_sid, existing_class = get_student_info(self.current_person_name)
        except Exception:
            existing_sid, existing_class = (None, None)
        existing_sid = existing_sid.strip() if isinstance(existing_sid, str) else existing_sid
        existing_class = existing_class.strip() if isinstance(existing_class, str) else existing_class
        self.current_student_id = student_id or existing_sid
        self.current_class = class_name or existing_class
        
        # Tạo thư mục cho người này: tên_masv
        suffix_id = (self.current_student_id or "unknown").replace(" ", "")
        self.current_folder_name = f"{self.current_person_name}_{suffix_id}"
        person_dir = os.path.join(DATASET_DIR, self.current_folder_name)
        folder_existed = os.path.exists(person_dir)
        os.makedirs(person_dir, exist_ok=True)
        self.collection_person_dir = person_dir
        self.collection_created_dir = not folder_existed
        self.collection_saved_any = False
        
        self.log_message(f"Bắt đầu thu thập {num_images} ảnh cho {person_name}", self.collection_log)
        self.log_message("Hệ thống sẽ TỰ ĐỘNG chụp ảnh khi phát hiện khuôn mặt", self.collection_log)

        try:
            # Chuẩn bị known encodings để chống trùng khuôn mặt (loại trừ đúng thư mục hiện tại)
            self.load_known_encodings(exclude_name=self.current_folder_name)
            self.collection_thread = CollectionCameraThread(CAMERA_INDEX, self.face_detector, auto_capture_mode=True)
            self.collection_thread.frame_ready.connect(self.update_collection_frame)
            self.collection_thread.auto_capture.connect(self.auto_capture_image)  # Kết nối signal tự động chụp
            self.collection_thread.error_occurred.connect(self.handle_camera_error)
            self.collection_thread.start()

            # Cập nhật nút: trái -> Dừng Thu Thập; phải -> Tạm Dừng
            try:
                self.btn_start_collection.clicked.disconnect()
            except Exception:
                pass
            self.btn_start_collection.setText("Dừng Thu Thập")
            self.btn_start_collection.setStyleSheet(self.get_button_style("#e74c3c"))
            self.btn_start_collection.clicked.connect(self.stop_collection)

            self.btn_stop_collection.setEnabled(True)
            self.btn_stop_collection.setText("Tạm Dừng")
            self.btn_stop_collection.setStyleSheet(self.get_button_style("#f39c12"))
            if self.collection_progress:
                self.collection_progress.setMaximum(num_images)
                self.collection_progress.setValue(0)
        except Exception as e:
            self.handle_camera_error(f"Lỗi: {str(e)}")
    
    def stop_collection(self):
        """
        Dừng thu thập
        """
        if self.collection_thread:
            self.collection_thread.stop()
            self.collection_thread = None
        
        self.collection_video_label.clear()
        self.collection_video_label.setText("Camera đã dừng")
        
        self.log_message(f"Đã dừng thu thập. Đã thu thập {self.collected_images} ảnh", self.collection_log)
        saved_any = self.collection_saved_any
        self._cleanup_collection_storage()
        # Reset lại nút
        try:
            self.btn_start_collection.clicked.disconnect()
        except Exception:
            pass
        self.btn_start_collection.setText("Bắt Đầu Thu Thập")
        self.btn_start_collection.setStyleSheet(self.get_button_style("#3498db"))
        self.btn_start_collection.clicked.connect(self.start_collection)
        self.btn_start_collection.setEnabled(True)

        self.btn_stop_collection.setEnabled(False)
        self.btn_stop_collection.setText("Tạm Dừng")
        self.btn_stop_collection.setStyleSheet(self.get_button_style("#f39c12"))
        if saved_any:
            if hasattr(self, 'input_person_name') and self.input_person_name:
                self.input_person_name.clear()
            if hasattr(self, 'input_student_id') and self.input_student_id:
                self.input_student_id.clear()
            if hasattr(self, 'input_class') and self.input_class:
                self.input_class.clear()
        self.current_person_name = ""
        self.collected_images = 0

    def pause_or_resume_collection(self):
        """
        Tạm dừng / Tiếp tục thu thập
        """
        if not self.collection_thread:
            return
        try:
            if not getattr(self.collection_thread, 'paused', False):
                # Tạm dừng
                if hasattr(self.collection_thread, 'pause'):
                    self.collection_thread.pause()
                self.btn_stop_collection.setText("Tiếp Tục")
                self.btn_stop_collection.setStyleSheet(self.get_button_style("#27ae60"))
                self.show_toast("Đã tạm dừng thu thập", success=False)
            else:
                # Tiếp tục
                if hasattr(self.collection_thread, 'resume'):
                    self.collection_thread.resume()
                self.btn_stop_collection.setText("Tạm Dừng")
                self.btn_stop_collection.setStyleSheet(self.get_button_style("#f39c12"))
                self.show_toast("Tiếp tục thu thập", success=True)
        except Exception as e:
            self.handle_camera_error(f"Lỗi khi tạm dừng/tiếp tục: {str(e)}")
    
    def _pause_collection_stream(self):
        """
        Tạm dừng luồng camera nội bộ (dùng khi xuất hiện cảnh báo).
        """
        try:
            if self.collection_thread and hasattr(self.collection_thread, 'pause'):
                self.collection_thread.pause()
        except Exception:
            pass
        # Không thay đổi trạng thái nút ở đây; stop_collection sẽ reset phù hợp.
    
    def _cleanup_collection_storage(self):
        """
        Xóa thư mục thu thập nếu là thư mục mới và chưa lưu ảnh nào.
        """
        try:
            if self.collection_person_dir and not self.collection_saved_any and self.collection_created_dir:
                if os.path.isdir(self.collection_person_dir):
                    shutil.rmtree(self.collection_person_dir)
        except Exception as e:
            self.log_message(f"Cảnh báo: không thể xóa thư mục thu thập tạm thời: {str(e)}", None)
        finally:
            self.collection_person_dir = None
            self.collection_created_dir = False
            self.collection_saved_any = False
    
    # BỎ HÀM HUẤN LUYỆN - KHÔNG CẦN NỮA
    # def start_training(self):
    #     pass
    
    def refresh_attendance(self):
        """
        Làm mới danh sách điểm danh
        """
        try:
            # Nếu bảng chưa sẵn sàng (chưa vào tab danh sách) thì bỏ qua lặng lẽ
            if not hasattr(self, 'attendance_table') or self.attendance_table is None:
                return
            date_str = None
            if hasattr(self, 'attendance_date_edit') and self.attendance_date_edit:
                qd = self.attendance_date_edit.date()
                py_dt = datetime(qd.year(), qd.month(), qd.day())
                date_str = py_dt.strftime(ATTENDANCE_DATE_FORMAT)
            records = load_attendance_records(date_str)
            # Áp dụng lọc theo tên, mã sv, lớp
            name_kw = (self.search_name.text().strip().lower() if hasattr(self, 'search_name') and self.search_name else '')
            sid_kw = (self.search_sid.text().strip().lower() if hasattr(self, 'search_sid') and self.search_sid else '')
            class_kw = (self.search_class.text().strip().lower() if hasattr(self, 'search_class') and self.search_class else '')
            if any([name_kw, sid_kw, class_kw]):
                def match(rec):
                    n = str(rec.get('Tên', '')).lower()
                    sid = str(rec.get('Mã SV', '')).lower()
                    cls = str(rec.get('Lớp', '')).lower()
                    return (name_kw in n) and (sid_kw in sid) and (class_kw in cls)
                records = [r for r in records if match(r)]
            
            # Cập nhật table: [Mã SV, Tên, Lớp, Ngày, Thời gian, Độ tin cậy]
            self.attendance_table.setRowCount(len(records))
            for i, record in enumerate(records):
                self.attendance_table.setItem(i, 0, QTableWidgetItem(record.get('Mã SV', '')))
                self.attendance_table.setItem(i, 1, QTableWidgetItem(record.get('Tên', '')))
                self.attendance_table.setItem(i, 2, QTableWidgetItem(record.get('Lớp', '')))
                self.attendance_table.setItem(i, 3, QTableWidgetItem(record.get('Ngày', '')))
                self.attendance_table.setItem(i, 4, QTableWidgetItem(record.get('Thời gian', '')))
                self.attendance_table.setItem(i, 5, QTableWidgetItem(record.get('Độ tin cậy', '')))
            
            # Cập nhật summary
            unique_persons = len(set(r.get('Tên', '') for r in records))
            summary_text = f"Tổng: {len(records)} lượt | Số người: {unique_persons}"
            if hasattr(self, 'attendance_summary') and self.attendance_summary:
                self.attendance_summary.setText(summary_text)
            
            if date_str:
                self.statusBar().showMessage(f"Đã làm mới danh sách điểm danh cho ngày {date_str}")
            else:
                self.statusBar().showMessage("Đã làm mới danh sách điểm danh")
        
        except Exception as e:
            # Không hiển thị popup; ghi log nhẹ trên status bar
            try:
                self.statusBar().showMessage(f"Không thể load dữ liệu điểm danh")
            except Exception:
                pass
            self.log_message(f"Lỗi load dữ liệu điểm danh: {str(e)}", None)
    
    def export_attendance(self):
        """
        Xuất file điểm danh (hỗ trợ cả Excel và CSV)
        """
        try:
            # Lấy ngày được chọn (nếu có)
            date_str = None
            if hasattr(self, 'attendance_date_edit') and self.attendance_date_edit:
                qd = self.attendance_date_edit.date()
                py_dt = datetime(qd.year(), qd.month(), qd.day())
                date_str = py_dt.strftime(ATTENDANCE_DATE_FORMAT)
            
            # Lấy dữ liệu điểm danh
            records = load_attendance_records(date_str)
            
            if not records:
                QMessageBox.information(self, "Thông báo", "Chưa có dữ liệu điểm danh!")
                return
            
            # Tạo tên file mặc định
            base_name = f"attendance_{date_str}" if date_str else "attendance_export"
            
            # Hiển thị hộp thoại lưu file với các lựa chọn định dạng
            file_path, selected_filter = QFileDialog.getSaveFileName(
                self, 
                "Xuất File", 
                base_name,
                "Excel Files (*.xlsx);;CSV Files (*.csv)",
                options=QFileDialog.Options()
            )
            
            if not file_path:
                return  # Người dùng đã hủy
                
            # Xác định định dạng file dựa trên phần mở rộng hoặc bộ lọc đã chọn
            is_excel = file_path.lower().endswith('.xlsx') or '*.xlsx' in selected_filter
            
            # Tạo DataFrame từ dữ liệu
            df = pd.DataFrame(records)
            
            # Xuất file tương ứng
            if is_excel:
                # Đảm bảo phần mở rộng là .xlsx
                if not file_path.lower().endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Xuất ra Excel
                df.to_excel(file_path, index=False, engine='openpyxl')
                
                # Định dạng file Excel
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import Font, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # Định dạng header
                    header_font = Font(bold=True, color='FFFFFF')
                    header_fill = '4472C4'  # Màu xanh dương
                    
                    # Định dạng border
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Định dạng cột header
                    for col in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = openpyxl.styles.PatternFill(start_color=header_fill, 
                                                             end_color=header_fill, 
                                                             fill_type='solid')
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Điều chỉnh độ rộng cột
                        column_letter = get_column_letter(col)
                        ws.column_dimensions[column_letter].width = 20
                    
                    # Định dạng các ô dữ liệu
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Lưu lại file Excel
                    wb.save(file_path)
                    
                except Exception as e:
                    print(f"Cảnh báo: Không thể định dạng file Excel: {str(e)}")
                
            else:  # Xuất CSV
                # Đảm bảo phần mở rộng là .csv
                if not file_path.lower().endswith('.csv'):
                    file_path += '.csv'
                
                # Xuất ra CSV với encoding UTF-8 có BOM để hiển thị đúng tiếng Việt
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
            
            # Thông báo thành công
            QMessageBox.information(
                self, 
                "Xuất file thành công", 
                f"Đã xuất dữ liệu điểm danh thành công!\n\nĐường dẫn: {file_path}"
            )
            
            # Hiển thị thông báo trên status bar
            if date_str:
                self.statusBar().showMessage(f"Đã xuất file điểm danh ngày {date_str}")
            else:
                self.statusBar().showMessage("Đã xuất dữ liệu điểm danh")
        
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Lỗi khi xuất file", 
                f"Đã xảy ra lỗi khi xuất file:\n{str(e)}"
            )
    
    def update_camera_frame(self, qt_image):
        """
        Cập nhật frame camera lên giao diện
        """
        pixmap = QPixmap.fromImage(qt_image)
        self.video_label.setPixmap(pixmap.scaled(
            self.video_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
        ))
        now = time.time()
        if self._last_fps_ts is None:
            self._last_fps_ts = now
        else:
            dt = now - self._last_fps_ts
            if dt > 0:
                fps = 1.0 / dt
                # Smooth
                self._fps = (self._fps * 0.8) + (fps * 0.2) if self._fps > 0 else fps
                if self.fps_label:
                    self.fps_label.setText(f"FPS: {self._fps:.1f}")
            self._last_fps_ts = now
    
    def update_collection_frame(self, qt_image, bgr_frame):
        """
        Cập nhật frame thu thập dữ liệu
        """
        self.current_frame = bgr_frame
        pixmap = QPixmap.fromImage(qt_image)
        self.collection_video_label.setPixmap(pixmap.scaled(
            self.collection_video_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
        ))
    
    def handle_camera_error(self, error_message):
        """
        Xử lý lỗi camera
        """
        QMessageBox.critical(self, "Lỗi Camera", error_message)
        self.log_message(f"✗ {error_message}", None)
        self.stop_camera()
        self.stop_collection()
    
    def keyPressEvent(self, event):
        """
        Xử lý sự kiện phím
        """
        # Nếu đang thu thập và nhấn SPACE (chụp thủ công)
        if event.key() == Qt.Key_Space and self.collection_thread and self.current_frame is not None:
            self.capture_image()
    
    def auto_capture_image(self, frame):
        """
        Tự động chụp ảnh khi phát hiện khuôn mặt
        """
        if frame is None or not self.current_person_name:
            return
        
        num_images = self.input_num_images.value()
        if self.collected_images >= num_images:
            self.log_message("Đã đủ số ảnh cần thu thập!", self.collection_log)
            # Tự động dừng thu thập ngay khi đủ số ảnh
            self.stop_collection()
            return 
        
        try:
            # Phát hiện khuôn mặt
            faces = self.face_detector.detect_faces(frame)
            
            if len(faces) == 0:
                return
            
            # Lấy khuôn mặt đầu tiên
            (x, y, w, h) = faces[0]
            face_img = frame[y:y+h, x:x+w]
            
            # Resize
            face_img = cv2.resize(face_img, IMAGE_SIZE)
            
            # Kiểm tra trùng khuôn mặt nếu có FaceEncoder (ngưỡng 70%)
            is_dup, dup_name, similarity = self.is_duplicate_face(face_img)
            stop_after_duplicate = False
            if is_dup:
                # Dừng camera ngay khi hiển thị cảnh báo
                self._pause_collection_stream()
                # Hiện dialog xác nhận
                msg = f"⚠️ Phát hiện khuôn mặt giống {dup_name or 'người đã có'} ({similarity:.1%})\n\n"
                msg += f"Sinh viên mới: {self.current_person_name}\n"
                msg += f"Sinh viên trùng: {dup_name or 'N/A'}\n\n"
                msg += "Bạn có muốn tiếp tục lưu ảnh này không?"
                reply = QMessageBox.question(
                    self, "Cảnh báo trùng lặp khuôn mặt", msg,
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if reply == QMessageBox.No:
                    self.show_toast(f"Đã hủy lưu do trùng với {dup_name or 'người đã có'}", success=False)
                    self.stop_collection()
                    return
                # Nếu chọn Yes, tiếp tục lưu và dừng camera sau khi lưu
                stop_after_duplicate = True

            # Lưu ảnh - Sử dụng PIL để xử lý Unicode path
            person_dir = os.path.join(DATASET_DIR, self.current_folder_name or self.current_person_name)
            os.makedirs(person_dir, exist_ok=True)  # Đảm bảo folder tồn tại
            
            suffix_id = (self.current_student_id or "unknown").replace(" ", "")
            img_path = os.path.join(person_dir, f"{self.current_person_name}_{suffix_id}_{self.collected_images}.jpg")
            
            # Chuyển BGR sang RGB cho PIL
            from PIL import Image
            face_img_rgb = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)
            pil_image = Image.fromarray(face_img_rgb)
            
            # Lưu bằng PIL (hỗ trợ Unicode path tốt hơn)
            try:
                pil_image.save(img_path, 'JPEG')
                self.collected_images += 1
                self.collection_saved_any = True
                self.log_message(f"✓ Tự động chụp ảnh {self.collected_images}/{num_images}", self.collection_log)
                self.log_message(f"  Đã lưu: {img_path}", self.collection_log)
                # Sau khi lưu ảnh thành công lần đầu: ghi thông tin sinh viên vào DB (nếu chưa có)
                try:
                    upsert_student_info(self.current_person_name, (self.current_student_id or None), (self.current_class or None))
                except Exception:
                    pass
                if self.collection_progress:
                    self.collection_progress.setValue(self.collected_images)
                # Toast thành công mỗi lần chụp
                self.show_toast(f"Đã chụp {self.collected_images}/{num_images} ảnh", success=True)
                # Mã hóa và lưu encoding vào DB để lần sau so khớp từ DB
                try:
                    if FaceEncoder is not None:
                        enc = FaceEncoder().encode_face(face_img)
                        if enc is not None:
                            save_student_encoding(self.current_person_name, enc)
                            # Cập nhật bộ known encodings trong phiên hiện tại
                            self.load_known_encodings(exclude_name=self.current_person_name)
                except Exception:
                    pass
            except Exception as e:
                self.log_message(f"✗ Lỗi lưu ảnh: {str(e)}", self.collection_log)
                # Toast thất bại khi lưu lỗi
                self.show_toast("Lỗi lưu ảnh khi thu thập", success=False)
                return
            
            # Kiểm tra xem đã đủ chưa
            if self.collected_images >= num_images:
                self.log_message(f"✓ Hoàn thành thu thập {self.collected_images} ảnh!", self.collection_log)
                # Tự động dừng thu thập, không hỏi
                # Toast hoàn thành thu thập
                self.show_toast(f"Hoàn thành thu thập {self.collected_images} ảnh", success=True)
                self.stop_collection()
                return

            if stop_after_duplicate:
                self.stop_collection()
                return
        
        except Exception as e:
            self.log_message(f"✗ Lỗi khi tự động chụp ảnh: {str(e)}", self.collection_log)
    
    def capture_image(self):
        """
        Chụp và lưu ảnh khuôn mặt
        """
        if self.current_frame is None or not self.current_person_name:
            return
        
        num_images = self.input_num_images.value()
        if self.collected_images >= num_images:
            self.log_message("Đã đủ số ảnh cần thu thập!", self.collection_log)
            # Tự động dừng thu thập ngay khi đủ số ảnh
            self.stop_collection()
            return
        
        try:
            # Phát hiện khuôn mặt
            faces = self.face_detector.detect_faces(self.current_frame)
            
            if len(faces) == 0:
                self.log_message("✗ Không phát hiện khuôn mặt", self.collection_log)
                return
            
            # Lấy khuôn mặt đầu tiên
            (x, y, w, h) = faces[0]
            face_img = self.current_frame[y:y+h, x:x+w]
            
            # Resize
            face_img = cv2.resize(face_img, IMAGE_SIZE)
            
            # Kiểm tra trùng khuôn mặt nếu có FaceEncoder (ngưỡng 70%)
            is_dup, dup_name, similarity = self.is_duplicate_face(face_img)
            stop_after_duplicate = False
            if is_dup:
                # Dừng camera ngay khi hiển thị cảnh báo
                self._pause_collection_stream()
                # Hiện dialog xác nhận
                msg = f"⚠️ Phát hiện khuôn mặt giống {dup_name or 'người đã có'} ({similarity:.1%})\n\n"
                msg += f"Sinh viên mới: {self.current_person_name}\n"
                msg += f"Sinh viên trùng: {dup_name or 'N/A'}\n\n"
                msg += "Bạn có muốn tiếp tục lưu ảnh này không?"
                reply = QMessageBox.question(
                    self, "Cảnh báo trùng lặp khuôn mặt", msg,
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if reply == QMessageBox.No:
                    self.show_toast(f"Đã hủy lưu do trùng với {dup_name or 'người đã có'}", success=False)
                    self.stop_collection()
                    return
                # Nếu chọn Yes, tiếp tục lưu nhưng sẽ dừng camera ngay sau đó
                stop_after_duplicate = True

            # Lưu ảnh - Sử dụng PIL để xử lý Unicode path
            person_dir = os.path.join(DATASET_DIR, (self.current_folder_name or self.current_person_name))
            os.makedirs(person_dir, exist_ok=True)  # Đảm bảo folder tồn tại
            
            suffix_id = (self.current_student_id or "unknown").replace(" ", "")
            img_path = os.path.join(person_dir, f"{self.current_person_name}_{suffix_id}_{self.collected_images}.jpg")
            
            # Chuyển BGR sang RGB cho PIL
            from PIL import Image
            face_img_rgb = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)
            pil_image = Image.fromarray(face_img_rgb)
            
            # Lưu bằng PIL (hỗ trợ Unicode path tốt hơn)
            try:
                pil_image.save(img_path, 'JPEG')
                self.collected_images += 1
                self.collection_saved_any = True
                self.log_message(f"✓ Đã chụp ảnh {self.collected_images}/{num_images}", self.collection_log)
                self.log_message(f"  Đã lưu: {img_path}", self.collection_log)
                if self.collection_progress:
                    self.collection_progress.setValue(self.collected_images)
                self.show_toast(f"Đã chụp {self.collected_images}/{num_images} ảnh", success=True)
                # Mã hóa và lưu encoding vào DB để lần sau so khớp từ DB
                try:
                    if FaceEncoder is not None:
                        enc = FaceEncoder().encode_face(face_img)
                        if enc is not None:
                            save_student_encoding(self.current_person_name, enc)
                            self.load_known_encodings(exclude_name=self.current_person_name)
                except Exception:
                    pass
            except Exception as e:
                self.log_message(f"✗ Lỗi lưu ảnh: {str(e)}", self.collection_log)
                self.show_toast("Lỗi lưu ảnh khi thu thập", success=False)
                return
            
            # Kiểm tra xem đã đủ chưa
            if self.collected_images >= num_images:
                self.log_message(f"✓ Hoàn thành thu thập {self.collected_images} ảnh!", self.collection_log)
                # Thông báo kèm Tên + Mã SV (đúng ngữ nghĩa)
                suffix_id = (self.current_student_id or "").strip()
                display_name = f"{self.current_person_name}{(' - ' + suffix_id) if suffix_id else ''}"
                self.show_toast(f"Thu thập thông tin của sinh viên {display_name} thành công", success=True)
                # Tự động dừng thu thập, không hỏi
                self.stop_collection()
                return

            if stop_after_duplicate:
                self.stop_collection()
                return
        
        except Exception as e:
            self.log_message(f"✗ Lỗi khi chụp ảnh: {str(e)}", self.collection_log)
    
    def log_message(self, message, log_widget):
        """
        Ghi log với timestamp ra widget tương ứng và console
        """
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_text = f"[{timestamp}] {message}"
        if log_widget:
            log_widget.append(log_text)
        print(log_text)

    def on_face_recognized(self, name, confidence):
        """
        Xử lý khi nhận diện được khuôn mặt
        """
        self.log_message(f"Nhận diện: {name} ({confidence:.2%})", self.recognition_log)
        if self.recognition_status_label:
            self.recognition_status_label.setText(f"{name} • {confidence:.0%}")
    
    def on_attendance_marked(self, name, time):
        """
        Xử lý khi điểm danh thành công
        """
        self.log_message(f"✓ Điểm danh: {name} - {time}", self.recognition_log)
        # Lấy mã SV để hiển thị
        sid = ''
        try:
            info = get_student_info(name)
            sid = info[0] if info and len(info) > 0 else ''
        except Exception:
            sid = ''
        display_name = f"{name}{(' - ' + sid) if sid else ''}"
        self.statusBar().showMessage(f"Đã điểm danh cho {display_name}")
        # Toast thành công điểm danh kèm Tên + Mã SV (đúng ngữ nghĩa)
        self.show_toast(f"{display_name} điểm danh thành công", success=True)

        # Refresh attendance list
        self.refresh_attendance()
        # Sau khi điểm danh: chỉ dừng camera, KHÔNG đóng ứng dụng
        self.stop_camera()

    def handle_recognition_failed(self):
        """
        Nhận diện thất bại sau thời gian yêu cầu: tự động dừng camera
        """
        self.log_message("✗ Nhận diện thất bại sau thời gian yêu cầu.", self.recognition_log)
        # Toast thất bại nhận diện
        self.show_toast("Nhận diện thất bại", success=False)
        # Không tự tắt camera khi thất bại; người dùng có thể tiếp tục hoặc tự bấm Tắt Camera
    
    def load_known_encodings(self, exclude_name=None):
        """
        Load encodings từ DB nếu có; nếu DB trống thì fallback quét dataset.
        """
        self.known_encodings = []
        self.known_names = []
        self.known_phashes = []
        # Ưu tiên nạp từ DB (loại trừ theo tên người hiện tại)
        try:
            exclude_name_db = self.current_person_name if exclude_name else None
            db_items = load_student_encodings(exclude_name=exclude_name_db)
            if db_items:
                for n, vec in db_items:
                    self.known_names.append(n)
                    self.known_encodings.append(vec)
                # Không return; tiếp tục nạp pHash để kiểm tra song song
        except Exception:
            pass
        # Quét dataset để bổ sung pHash và (nếu có) encodings
        encoder = None
        if FaceEncoder is not None:
            try:
                encoder = FaceEncoder()
            except Exception:
                encoder = None
        try:
            persons = [d for d in os.listdir(DATASET_DIR) if os.path.isdir(os.path.join(DATASET_DIR, d))]
            for person in persons:
                if exclude_name and person == exclude_name:
                    continue
                person_dir = os.path.join(DATASET_DIR, person)
                for f in os.listdir(person_dir):
                    if not f.lower().endswith((".jpg", ".jpeg", ".png", ".bmp")):
                        continue
                    img_path = os.path.join(person_dir, f)
                    try:
                        # Dùng PIL để đọc ảnh an toàn Unicode
                        from PIL import Image
                        try:
                            pil_img = Image.open(img_path).convert('RGB')
                        except Exception:
                            continue
                        img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
                        if img is None or img.size == 0:
                            continue
                        # pHash luôn được tính
                        ph = self._compute_phash(img)
                        if ph is not None:
                            self.known_phashes.append(ph)
                            self.known_names.append(person)
                        # Nếu có encoder: bổ sung vector đặc trưng
                        if encoder is not None:
                            enc = encoder.encode_face(img)
                            if enc is not None:
                                self.known_encodings.append(enc)
                    except Exception:
                        continue
        except Exception:
            pass

    def is_duplicate_face(self, face_bgr):
        """
        Kiểm tra khuôn mặt có trùng với người đã có (>=70% similarity) không.
        Dùng cả cosine (FaceEncoder) và pHash; lấy điểm cao nhất.
        Trả về: (is_duplicate, best_name, best_similarity)
        """
        # Đảm bảo đã có dữ liệu so khớp
        if not getattr(self, 'known_encodings', None) and not getattr(self, 'known_phashes', None):
            self.load_known_encodings(exclude_name=(self.current_folder_name or self.current_person_name))
        import numpy as np
        best_name = None
        best_sim = 0.0
        # 1) pHash similarity
        try:
            target_ph = self._compute_phash(face_bgr)
            if target_ph is not None and getattr(self, 'known_phashes', None):
                def hamming_sim(a, b):
                    a = np.asarray(a, dtype=np.uint8)
                    b = np.asarray(b, dtype=np.uint8)
                    dist = int(np.count_nonzero(a != b))
                    return 1.0 - dist / 64.0
                ph_sims = [hamming_sim(target_ph, k) for k in self.known_phashes]
                if ph_sims:
                    idx = int(np.argmax(ph_sims))
                    if ph_sims[idx] > best_sim:
                        best_sim = float(ph_sims[idx])
                        # Map idx theo known_names nếu còn phạm vi
                        if 0 <= idx < len(self.known_names):
                            best_name = self.known_names[idx]
        except Exception:
            pass
        # 2) Cosine similarity (FaceEncoder)
        try:
            if FaceEncoder is not None:
                enc = FaceEncoder().encode_face(face_bgr)
                if enc is not None and getattr(self, 'known_encodings', None):
                    def cos_sim(a, b):
                        a = np.asarray(a); b = np.asarray(b)
                        denom = (np.linalg.norm(a) * np.linalg.norm(b))
                        if denom == 0:
                            return 0.0
                        return float(np.dot(a, b) / denom)
                    cs_sims = [cos_sim(enc, k) for k in self.known_encodings]
                    if cs_sims:
                        idx = int(np.argmax(cs_sims))
                        if cs_sims[idx] > best_sim:
                            best_sim = float(cs_sims[idx])
                            # Không có mảng tên riêng cho encodings, dùng known_names nếu có cùng thứ tự
                            if 0 <= idx < len(self.known_names):
                                best_name = self.known_names[idx]
        except Exception:
            pass
        return (best_sim >= DUPLICATE_SIMILARITY_THRESHOLD, best_name, best_sim)

    def _compute_phash(self, bgr_img):
        """Tính pHash 64-bit (trả về mảng 64 phần tử 0/1) cho ảnh BGR."""
        try:
            import numpy as np
            gray = cv2.cvtColor(bgr_img, cv2.COLOR_BGR2GRAY)
            # Resize lớn hơn trước khi DCT để ổn định
            small = cv2.resize(gray, (32, 32), interpolation=cv2.INTER_AREA)
            small = np.float32(small)
            dct = cv2.dct(small)
            dct_low = dct[:8, :8]
            mean = dct_low.mean()
            bits = (dct_low > mean).astype(np.uint8).flatten()
            return bits  # 64 phần tử
        except Exception:
            return None

    # BỎ CÁC HÀM LIÊN QUAN ĐẾN HUẤN LUYỆN
    # def on_training_progress(self, message):
    #     pass
    # 
    # def on_training_complete(self, success, message):
    #     pass
    
    def reload_recognition_data(self):
        """
        Reload dữ liệu nhận diện (khi có người dùng mới)
        """
        if not self.recognition_thread:
            QMessageBox.warning(self, "Lỗi", "Hệ thống nhận diện chưa khởi động!")
            return
        
        self.log_message("🔄 Đang reload dữ liệu người dùng...", self.recognition_log)
        
        try:
            # Reload templates cho SimpleRecognitionThread
            if hasattr(self.recognition_thread, 'reload_templates'):
                self.recognition_thread.reload_templates()
                self.log_message("✓ Đã reload templates thành công!", self.recognition_log)
                QMessageBox.information(
                    self, "Thành công",
                    "Đã cập nhật dữ liệu người dùng mới!\n\n"
                    "Hệ thống sẽ nhận diện người dùng mới ngay lập tức."
                )
            # Reload model cho RecognitionThread (nếu dùng)
            elif hasattr(self.recognition_thread, 'reload_model'):
                self.recognition_thread.reload_model()
                self.log_message("✓ Đã reload model thành công!", self.recognition_log)
                QMessageBox.information(
                    self, "Thành công",
                    "Đã cập nhật mô hình nhận diện!\n\n"
                    "Lưu ý: Cần huấn luyện lại mô hình để nhận diện người mới."
                )
            else:
                self.log_message("✗ Thread không hỗ trợ reload", self.recognition_log)
                QMessageBox.warning(
                    self, "Lỗi",
                    "Thread hiện tại không hỗ trợ reload!\n"
                    "Vui lòng khởi động lại hệ thống."
                )
        except Exception as e:
            self.log_message(f"✗ Lỗi reload: {str(e)}", self.recognition_log)
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi reload dữ liệu:\n{str(e)}")

    def apply_styles(self):
        qss = """
        QWidget { font-family: 'Segoe UI', Arial; font-size: 13px; color: #2c3e50; }
        QMainWindow { background: #f5f7fb; }

        /* Tabs */
        QTabWidget::pane { border: 1px solid #e3e8f0; border-radius: 10px; background: #ffffff; }
        QTabBar::tab { background: #eef2f7; padding: 10px 18px; margin: 4px; border-radius: 8px; color: #34495e; }
        QTabBar::tab:selected { background: #3498db; color: #ffffff; }

        /* Buttons */
        QPushButton { background-color: #3498db; color: white; border: none; padding: 10px 14px; font-size: 14px; font-weight: 600; border-radius: 8px; }
        QPushButton:hover { background-color: #2d86c4; }
        QPushButton:disabled { background-color: #b8c4ce; color: #f0f3f6; }

        /* Inputs */
        QLineEdit, QSpinBox { background: #ffffff; border: 1px solid #dfe6ee; border-radius: 8px; padding: 8px; }
        QLineEdit:focus, QSpinBox:focus { border: 1px solid #3498db; }
        QCheckBox { padding: 4px; }

        /* Status pill */
        QLabel#StatusPill { padding: 6px 10px; border-radius: 14px; background: #eaf5ff; color: #2c3e50; font-weight: 600; }

        /* Table */
        QTableWidget { background: #ffffff; border: 1px solid #e3e8f0; border-radius: 10px; gridline-color: #eef2f7; }
        QHeaderView::section { background: #f6f9fc; padding: 8px; border: none; border-right: 1px solid #e9eef5; font-weight: 600; color: #34495e; }
        QTableWidget::item { padding: 8px; }
        QTableWidget::item:selected { background: #d7ecff; color: #1b2b40; }

        /* Progress */
        QProgressBar { background: #eef3f8; border: 1px solid #e3e8f0; border-radius: 8px; padding: 3px; }
        QProgressBar::chunk { background-color: #27ae60; border-radius: 6px; }
        """
        self.setStyleSheet(qss)
    
    def closeEvent(self, event):
        """
        Xử lý khi đóng cửa sổ
        """
        # Dừng tất cả threads
        if self.camera_thread:
            self.camera_thread.stop()
        if self.collection_thread:
            self.collection_thread.stop()
        if self.recognition_thread:
            self.recognition_thread.stop()
        if hasattr(self, 'multi_recognition_thread') and self.multi_recognition_thread:
            self.multi_recognition_thread.stop()
        # Không còn training_thread nữa
        event.accept()
